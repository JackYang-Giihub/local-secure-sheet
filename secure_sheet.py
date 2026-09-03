#!/usr/bin/env python3
"""Local Excel/CSV column encryption tool. AES-256-GCM + PBKDF2-HMAC-SHA256."""
from __future__ import annotations
import base64, csv, json, os, secrets, sys, zipfile
from pathlib import Path

from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes
from openpyxl import load_workbook, Workbook

MAGIC = "LSS1"
ITERATIONS = 600_000

def key_from_password(password: str, salt: bytes) -> bytes:
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=ITERATIONS)
    return kdf.derive(password.encode("utf-8"))

def enc(value, key):
    if value is None or value == "": return value
    nonce = secrets.token_bytes(12)
    data = str(value).encode("utf-8")
    return "lss:" + base64.urlsafe_b64encode(nonce + AESGCM(key).encrypt(nonce, data, None)).decode()

def dec(value, key):
    if not isinstance(value, str) or not value.startswith("lss:"): return value
    raw = base64.urlsafe_b64decode(value[4:].encode())
    nonce, payload = raw[:12], raw[12:]
    return AESGCM(key).decrypt(nonce, payload, None).decode("utf-8")

def read_table(path: Path):
    if path.suffix.lower() == ".txt":
        return [["内容"]] + [[line.rstrip("\n\r")] for line in path.open("r", encoding="utf-8-sig")]
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f: return list(csv.reader(f))
    wb = load_workbook(path, data_only=False, keep_vba=path.suffix.lower()==".xlsm")
    ws = wb.active
    return [[c.value for c in row] for row in ws.iter_rows()]

def write_table(path: Path, rows):
    if path.suffix.lower() == ".txt":
        path.write_text("\n".join("" if r[0] is None else str(r[0]) for r in rows[1:]), encoding="utf-8")
        return
    if path.suffix.lower() == ".csv":
        with path.open("w", encoding="utf-8-sig", newline="") as f: csv.writer(f).writerows(rows)
        return
    wb = Workbook(); ws = wb.active
    for r, row in enumerate(rows, 1):
        for c, value in enumerate(row, 1): ws.cell(r, c).value = value
    wb.save(path)

def transform(input_file, output_file, columns, password, decrypt=False):
    rows = read_table(Path(input_file))
    if not rows: raise ValueError("文件为空")
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    missing = [c for c in columns if c not in headers]
    if missing: raise ValueError("找不到列：" + ", ".join(missing) + "；可用列：" + ", ".join(headers))
    if decrypt:
        meta = Path(str(input_file) + ".meta.json")
        if not meta.exists(): raise ValueError("解密需要与加密文件同目录的 .meta.json 文件")
        info = json.loads(meta.read_text(encoding="utf-8"))
        if info.get("magic") != MAGIC: raise ValueError("无效的加密参数文件")
        salt = base64.b64decode(info["salt"])
    else:
        salt = secrets.token_bytes(16)
    key = key_from_password(password, salt)
    for row in rows[1:]:
        for col in columns:
            i = headers.index(col)
            if i < len(row): row[i] = dec(row[i], key) if decrypt else enc(row[i], key)
    write_table(Path(output_file), rows)
    if not decrypt:
        Path(str(output_file) + ".meta.json").write_text(json.dumps({"magic":MAGIC,"salt":base64.b64encode(salt).decode(),"iterations":ITERATIONS}, ensure_ascii=False), encoding="utf-8")

def transform_many(files, output_dir, columns, password, decrypt=False):
    output_dir = Path(output_dir); output_dir.mkdir(parents=True, exist_ok=True)
    outputs = []
    for source in files:
        p = Path(source); marker = "_解密" if decrypt else "_加密"
        out = output_dir / f"{p.stem}{marker}{p.suffix}"
        transform(str(p), str(out), columns, password, decrypt)
        outputs.append(str(out))
    return outputs

def encrypt_text(value: str, password: str) -> str:
    salt = secrets.token_bytes(16); key = key_from_password(password, salt); nonce = secrets.token_bytes(12)
    payload = AESGCM(key).encrypt(nonce, value.encode("utf-8"), None)
    return "lss1:" + base64.urlsafe_b64encode(salt + nonce + payload).decode()

def decrypt_text(token: str, password: str) -> str:
    if not token.startswith("lss1:"): raise ValueError("不是有效的单条文本密文")
    raw = base64.urlsafe_b64decode(token[5:].encode()); salt, nonce, payload = raw[:16], raw[16:28], raw[28:]
    return AESGCM(key_from_password(password, salt)).decrypt(nonce, payload, None).decode("utf-8")

def main():
    if len(sys.argv) != 6: print("用法: python secure_sheet.py encrypt|decrypt 输入文件 输出文件 列名(逗号分隔) 密码"); return 2
    mode, inp, out, cols, password = sys.argv[1:]
    meta = Path(inp + ".meta.json") if mode == "decrypt" else None
    if meta and meta.exists():
        salt = base64.b64decode(json.loads(meta.read_text(encoding="utf-8"))["salt"])
        # decrypt uses the original salt stored beside the encrypted file
        rows = read_table(Path(inp)); headers=[str(x).strip() for x in rows[0]]; key=key_from_password(password,salt)
        for row in rows[1:]:
            for col in cols.split(','):
                i=headers.index(col.strip());
                if i < len(row): row[i]=dec(row[i],key)
        write_table(Path(out),rows)
    elif mode == "decrypt":
        raise ValueError("解密需要与加密文件同目录的 .meta.json 密钥参数文件")
    else:
        transform(inp,out,[x.strip() for x in cols.split(',')],password,False)
    print(f"完成：{out}")

if __name__ == '__main__': main()
